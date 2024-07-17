#!/usr/bin/python3

# Author: James O'Brien
# 03-05-2024

import os
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment

# Create a parser object
argparser = argparse.ArgumentParser(description='Search .log files for NBO values')
argparser.add_argument('-g', '--grep', nargs='+', required=True, help='NBO values to search for (eg. LP, BD*, etc.)')
argparser.add_argument('-v', '--exclude', nargs='+', help='Exclude NBO values to search for (eg. F, H, S, N)')
args = argparser.parse_args()

search_terms = args.grep
exclude_terms = args.exclude or []  # Set exclude_terms to an empty list if it's None

# Create a new workbook
workbook = openpyxl.Workbook()

# Iterate through all log files in the current working directory
for file in os.listdir('.'):
    if file.endswith('.log'):
        # Create a new worksheet for each log file
        worksheet = workbook.create_sheet(file[:-4])
        
        # Write the header row
        worksheet['A1'] = 'Donor NBO (i)'
        worksheet['B1'] = 'Acceptor NBO (j)'
        worksheet['C1'] = 'E(2) (kcal/mol)'
        worksheet['D1'] = 'E(j)-E(i) (a.u.)'
        worksheet['E1'] = 'F(i,j) (a.u.)'
        
        # Set the header font and alignment
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in worksheet['A1:E1'][0]:
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Search the log file for NBO values
        grep_command = f'grep "{search_terms[0]}" {file}'
        for search_term in search_terms[1:]:
            grep_command += f' | grep "{search_term}"'
        if exclude_terms:
            for exclude_term in exclude_terms:
                grep_command += f' | grep -v "{exclude_term}"'
        
        grep_output = os.popen(grep_command).read().strip()
        
        # Write the NBO values to the worksheet
        row = 2
        for line in grep_output.split('\n'):
            if line.strip():
                columns = line.split()
                if len(columns) >= 16 and all(col.replace('.', '', 1).isdigit() for col in columns[13:16]):
                    donor_nbo = ' '.join(columns[:6])
                    acceptor_nbo = ' '.join(columns[6:13])
                    e2 = float(columns[13])
                    e_diff = float(columns[14])
                    f_ij = float(columns[15])
                    
                    worksheet.cell(row=row, column=1, value=donor_nbo)
                    worksheet.cell(row=row, column=2, value=acceptor_nbo)
                    worksheet.cell(row=row, column=3, value=e2)
                    worksheet.cell(row=row, column=4, value=e_diff)
                    worksheet.cell(row=row, column=5, value=f_ij)
                    row += 1
                else:
                    print(f"Skipping line: {line}")
        
        # Autofit the columns
        for column in range(1, 6):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].auto_size = True

# Remove the default 'Sheet' worksheet
workbook.remove(workbook['Sheet'])

# Save the workbook to a file
workbook.save('nbo_values.xlsx')
